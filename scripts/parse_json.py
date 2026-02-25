"""
Parse ./tmp/data.json to generate json and spreadsheet.
"""

import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from utils.utils import parse_json_args, setup_logger

DATA_PATH = "./tmp/data.json"
CATEGORY_PATH = "./tmp/category.json"
ID_MAPPING_PATH = "./tmp/mapping.json"
OIL_XLSX_PATH = "./oils.xlsx"
OIL_JSON_PATH = "./oils.json"
RECIPE_JSON_PATH = "./recipes.json"
RECIPE_XLSX_PATH = "./recipes.xlsx"
FINAL_RESULT_PATH = "./results.json"

RECIPE_DATABASE_ID = "3425407372818098406"
I2_LANGUAGES_ID = "-4669794531358937986"

COLUMN_MAPPING = {
    "displayName": "Name",
    "includedInDemo": "Demo",
    "includedInEarlyAccess": "Early Access",
    "basePrice": "Base Price",
    "CostsDurability": "Costs Durability",
    "Kick": "Recoil",
    "Reload Speed": "Reload Speed",
    "Damage": "Damage",
    "Critical damage chance": "Crit Chance",
    "Disables aiming": "Disables Aiming",
    "Projectile drag multiplier": "Drag Mult",
    "Spread": "Spread",
    "Loot Chance Multiplier": "Loot Chance Multiplier",
    "Bullet bounces": "Bullet Bounces",
    "Time scale": "Bullet Speed",
    # "Damage%": "Damage%",
    "Bullet size": "Bullet Size",
    "Bullet drop": "Bullet Drop",
    "No money drops": "No Money Drops",
    "Move speed": "Move Speed",
    "Rounds per minute": "RPM",
    "Bullet Penetration": "Bullet Penetration",
    "Jump power": "Jump Power",
    "Max Durability": "Max Durability",
    "Number of projectiles": "Projectile Amount",
    "Projectile bounciness": "Bullet Bounciness",
    "Chance this consumes ammo": "Ammo Consume Chance",
    "Chance to consume extra ammo": "Consume Extra Ammo Chance",
    "No organs drop": "No Organs Drop",
    "Durability Per Shot": "Durability Per Shot",
    "Projectile force multiplier": "Projectile Force Multiplier",
    "Accuracy when moving": "Accuracy When Moving",
    "Enchantment Random Oil": "Enchantment Random Oil",
    "MISC": "MISC",  # Not a built-in one, for sheet name conversion
}

# ItemDescriptions/WeaponType_Pistol
WEAPON_TYPE = {
    1: "AssaultRifle",
    3: "LMG",
    # 4: "Melee",
    5: "Pistol",
    6: "Revolver",
    7: "Rifle",
    8: "Shotgun",
    9: "SMG",
    10: "Sniper",
    # 11: "Throwable",
}

# WorldResource/Resource_Ammo_12Ga_short
AMMO_TYPE = {
    # 0 for Melee
    1: "9mm",
    2: "12Ga",
    3: "556",
    4: "762",
    5: "50BMG",
    7: "EnergyCell",
}

DAMAGE_MULTIPLIERS = {
    "caliber": {
        1: (60, 1),
        2: (20, 8),
        3: (80, 1),
        4: (100, 1),
        5: (200, 1),
        7: (50, 1),
    },
    "weaponType": {1: 1.2, 3: 1.0, 5: 1.0, 6: 1.6, 7: 2.0, 8: 1.0, 9: 1.0, 10: 2.0},
}

EFFECT_TYPES = list(COLUMN_MAPPING.values())[4:]

args = parse_json_args()
logger = setup_logger(args.logging_level)
cnt = 0


def build_enchantment_object(type, item_id):
    item_data = data[item_id]
    global cnt
    cnt += 1
    logger.info(f"Parsing {cnt:>4} '%s'", item_data["m_Name"])
    results = {
        **get_basic_attributes(type, item_id),
        "includedInDemo": item_data["includedInDemo"],
        "includedInEarlyAccess": item_data["includedInEarlyAccess"],
    }

    item_definition = get_enchantment_definition(
        id_mapping["enchantmentDefinition"][
            str(item_data["appliesEnchantment"]["value"])
        ],
    )
    results |= item_definition

    if args.dev:
        results["artwork"] = str(item_data["artwork"]["m_PathID"])
        # results["definition"] = item_definition  # For structural parsing

    return results


def format_value(value):
    return float(f"{value:.2f}") if isinstance(value, float) else value


def get_enchantment_definition(oil_definition_id):
    # EnchantmentDefinition_*Oil
    definition_data = data[oil_definition_id]
    modifier_definitions = get_modifiers_definition(definition_data["modifiersApplied"])
    return {
        "CostsDurability": definition_data["CostsDurability"],
        **modifier_definitions,
        "displayFields": list(modifier_definitions),
    }


def get_modifiers_definition(modifiers):
    results = {}
    for modifier in modifiers:
        # itemDescriptionName and label not reliable, use m_Name instead
        # ItemAttributes/ConsumeAmmoChance_itemDescription
        # ItemAttributes/ConsumeAmmoChance_label
        name = data[id_mapping["attributeModifier"][str(modifier["attribute"])]][
            "m_Name"
        ]
        # 100: boolean/add, 200: multiplier, 300: bullet size
        # try:
        #     mod_type = modifier["modType"]
        # except KeyError:
        #     logger.warning("modType not found!")
        #     mod_type = 0
        value = format_value(modifier["value"])
        # if name == "Damage" and mod_type == 200:
        #     results["Damage%"] = value
        # else:
        results[name] = value
    return results


def get_oil_types(info):
    oil_types = list(info.keys())
    return oil_types if oil_types else ["MISC"]


def adjust_width(filename):
    wb = openpyxl.load_workbook(filename)
    for worksheet in wb.sheetnames:
        ws = wb[worksheet]
        dim_holder = DimensionHolder(worksheet=ws)
        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(
                ws, min=col, max=col, width=20
            )
        ws.column_dimensions = dim_holder
    wb.save(filename)


def dump_oil_xlsx(oil_infos, oil_groups):
    with pd.ExcelWriter(OIL_XLSX_PATH, engine="openpyxl") as writer:
        df = pd.DataFrame(oil_infos)
        df = df.rename(columns=COLUMN_MAPPING)
        df.to_excel(writer, sheet_name="Comparison Chart", index=False)
        for group_name, oil_group_infos in oil_groups.items():
            df = pd.DataFrame(oil_group_infos)
            df.to_excel(writer, sheet_name=group_name, index=False)
    adjust_width(OIL_XLSX_PATH)


def parse_enchantment_data(type, item_ids):
    item_infos = [build_enchantment_object(type, item_id) for item_id in item_ids]
    oil_groups = defaultdict(list)
    for oil_info in item_infos:
        for oil_type in get_oil_types(oil_info):
            oil_groups[oil_type].append(oil_info)

    if type == "oil":
        with open(OIL_JSON_PATH, "w", encoding="utf8") as f:
            json.dump({"all": item_infos} | oil_groups, f, ensure_ascii=False, indent=4)
        dump_oil_xlsx(item_infos, oil_groups)
    return item_infos


def dump_recipe_xlsx(recipe_infos, recipes_of_items):
    with pd.ExcelWriter(RECIPE_XLSX_PATH, engine="openpyxl") as writer:
        df = pd.DataFrame(recipe_infos)
        df.to_excel(writer, sheet_name="Recipes", index=False)
        df = pd.DataFrame.from_dict(recipes_of_items, orient="index")
        df.to_excel(writer, sheet_name="Recipes of Items")
    adjust_width(RECIPE_XLSX_PATH)


def get_recipe_mapping(src_data):
    mapping = {}
    for k, v in src_data.items():
        # Use three keys to identify an actual item, wtf perfect random
        if "displayName" in v and "id" in v and "artwork" in v:
            mapping[v["id"]["value"]] = k
    return mapping


def build_recipe_object(src_data, id_mapping, recipe_data):
    global cnt
    cnt += 1
    # Recipe doesn't have displayName
    logger.info(f"Parsing {cnt:>4} '%s'", recipe_data["name"])
    result = {
        "recipeName": recipe_data["name"],
        "name": src_data[id_mapping[recipe_data["createsItem"]["value"]]]["m_Name"],
        "quantity": recipe_data["quantityCreated"],
        "artwork": str(
            src_data[id_mapping[recipe_data["createsItem"]["value"]]]["artwork"][
                "m_PathID"
            ]
        ),
        "itemsNeeded": [],
        "newPrice": src_data[id_mapping[recipe_data["createsItem"]["value"]]]["basePrice"]
        * recipe_data["quantityCreated"],
    }

    old_price = 0
    for item_data in recipe_data["itemsNeeded"]:
        real_item_data = src_data[id_mapping[item_data["item"]["value"]]]
        item_name = real_item_data["m_Name"]

        try:
            result["itemsNeeded"].append(
                {
                    "name": item_name,
                    "quantity": item_data["quantity"],
                    "artwork": str(real_item_data["artwork"]["m_PathID"]),
                }
            )
            old_price += real_item_data["basePrice"] * item_data["quantity"]
            continue
        except KeyError:
            # Some items like cactus is not available now
            logger.warning(
                "Artwork id not found: '%s'", real_item_data["artwork"]["m_PathID"]
            )

    result["oldPrice"] = old_price
    return result


def parse_recipe_data(data):
    id_mapping = get_recipe_mapping(data)

    recipes = data[RECIPE_DATABASE_ID]["recipes"]
    recipe_infos = [
        build_recipe_object(data, id_mapping, recipe_data) for recipe_data in recipes
    ]
    recipes_of_items = defaultdict(list)
    for recipe_info in recipe_infos:
        recipes_of_items[recipe_info["name"]].append(
            {
                "quantity": recipe_info["quantity"],
                "itemsNeeded": recipe_info["itemsNeeded"],
            }
        )

    with open(RECIPE_JSON_PATH, "w", encoding="utf8") as f:
        # json.dump(
        #     {"all": recipe_infos} | recipes_of_items, f, ensure_ascii=False, indent=4
        # )
        json.dump(recipe_infos, f, ensure_ascii=False, indent=4)

    dump_recipe_xlsx(recipe_infos, recipes_of_items)
    return recipe_infos


def get_damage(
    caliber, weapon_type, damage_multiplier, ammo_multiplier, ammo_overwrite=None
):
    damage, ammo_per_shot = DAMAGE_MULTIPLIERS["caliber"][caliber]
    ammo_per_shot = ammo_overwrite if ammo_overwrite else ammo_per_shot
    #
    result = f"{int(damage * DAMAGE_MULTIPLIERS['weaponType'][weapon_type] * damage_multiplier)}"

    if ammo_per_shot > 1:
        result += f"x{ammo_per_shot}"
    if ammo_multiplier > 1:
        result += f"x{ammo_multiplier}"
    return result


def parse_weapon_data(type, item_ids):
    results = []
    global cnt
    for item_id in item_ids:
        item_data = data[item_id]
        cnt += 1
        # Might include throwables
        if "slotType" in item_data and item_data["slotType"] == 7:
            logger.info(f"Parsing {cnt:>4} '%s'", item_data["m_Name"])
            results.append(
                {
                    **get_basic_attributes(type, item_id),
                    "Durability": item_data["maxDurability"],
                    "MagSize": item_data["iAmmoMax"],
                    "ReloadSpeed": item_data["fReloadTime"],
                    "RPM": item_data["rpm"],
                    "Damage": get_damage(
                        item_data["caliber"],
                        item_data["weaponType"],
                        format_value(item_data["damageMultiplier"]),
                        item_data["iMaxAmmoPerShot"],
                        3 if item_data["m_Name"] == "Weapon_Augusta" else None,
                    ),
                    "ammoPerShot": item_data["iMaxAmmoPerShot"],
                    "weaponTypeMultiplier": DAMAGE_MULTIPLIERS["weaponType"][
                        item_data["weaponType"]
                    ],
                    "damageMultiplier": format_value(item_data["damageMultiplier"]),
                    "Type": WEAPON_TYPE[item_data["weaponType"]],
                    "AmmoType": AMMO_TYPE[item_data["caliber"]],
                    "spreadPerCaliber": {
                        d["Caliber"]: d["Spread"] for d in item_data["spreadPerCaliber"]
                    },
                    **get_modifiers_definition(item_data["baseAttributes"]),
                    "displayFields": [
                        "Type",
                        "AmmoType",
                        "Damage",
                        "RPM",
                        "MagSize",
                        "Spread",
                        "Durability",
                    ],
                }
            )

    return results


def get_i2languages(data):
    i2languages = data[I2_LANGUAGES_ID]["mSource"]["mTerms"]
    return {flavor["Term"]: flavor["Languages"] for flavor in i2languages}


def get_basic_attributes(type, item_id):
    item_data = data[item_id]
    return {
        "id": item_id,
        "m_Name": item_data["m_Name"],
        "type": type,
        "artwork": str(item_data["artwork"]["m_PathID"]),
        "basePrice": item_data["basePrice"],
        "itemQuality": item_data["itemQuality"],
    }


def parse_chamber_chisel(type, item_ids):
    results = []
    global cnt
    for item_id in item_ids:
        item_data = data[item_id]
        cnt += 1
        logger.info(f"Parsing {cnt:>4} '%s'", item_data["m_Name"])
        results.append(
            {
                **get_basic_attributes(type, item_id),
                # "modifiesCaliber": AMMO_TYPE[item_data["modifiesCaliber"]],
                "modifiesCaliber": AMMO_TYPE[item_data["modifiesCaliber"]],
                "displayFields": [],
            }
        )

    return results


def parse_item_data(category, results):
    parse_functions = {
        "weapon": parse_weapon_data,
        "chamberChisel": parse_chamber_chisel,
        "oil": parse_enchantment_data,
        "scroll": parse_enchantment_data,
        "muzzle": parse_attachment_data,
        "scope": parse_attachment_data,
        "laserSight": parse_attachment_data,
        "chamber": parse_attachment_data,
        "insurance": parse_attachment_data,
    }

    for k, v in category.items():
        if isinstance(v, dict):
            parse_item_data(v, results)
        else:  # v is list of ids
            # "useType" cannot tell oil and scroll apart, use self-defined useType instead
            results.extend(parse_functions[k](k, v))


def parse_attachment_data(type, item_ids):
    results = []
    global cnt
    for item_id in item_ids:
        item_data = data[item_id]
        cnt += 1
        logger.info(f"Parsing {cnt:>4} '%s'", item_data["m_Name"])

        modifier_definitions = get_modifiers_definition(
            item_data["modifiersOnAttachToItem"]
        )
        results.append(
            {
                **get_basic_attributes(type, item_id),
                **modifier_definitions,
                "displayFields": list(modifier_definitions),
            }
        )
    return results


def parse_i18n():
    data[I2_LANGUAGES_ID]["mSource"]["mLanguages"].pop()  # ar is not implemented yet
    i18n = {x["Code"]: {} for x in data[I2_LANGUAGES_ID]["mSource"]["mLanguages"]}
    languages = i18n.keys()
    print(
        "Languages:",
        {x["Code"]: x["Name"] for x in data[I2_LANGUAGES_ID]["mSource"]["mLanguages"]},
    )
    for k, v in i2languages.items():
        if (
            k.startswith("ItemDescriptions/")
            or (k.startswith("WorldResource/"))
            or k.startswith("Items/")
        ):
            for language, text in zip(languages, v):
                i18n[language][k] = text

        if k.startswith("ItemAttributes/") and k.endswith("_itemDescription"):
            # Try itemDescription first, use label value as backup
            if not v:
                v = i2languages[k.replace("_itemDescription", "_label")]

            if k == "ItemAttributes/ProjectilieFlameThrower_itemDescription":
                k = "ItemAttributes/ProjectileFlameThrower_itemDescription"
            elif k == "ItemAttributes/ProjectileOnHitChainLighting_itemDescription":
                k = "ItemAttributes/ProjectileOnHitChainLightning_itemDescription"
            elif (
                k == "ItemAttributes/EnchantmentIncreaseHeadshotDamage_itemDescription"
            ):
                k = "ItemAttributes/EnchantmentIncreasedHeadshotDamage_itemDescription"

            # Patch this dog shit
            for language, text in zip(languages, v):
                i18n[language][k] = text

    Path("locales").mkdir(parents=True, exist_ok=True)
    for language, content in i18n.items():
        with open(Path(f"locales/{language}.json"), "w", encoding="utf8") as f:
            json.dump(content, f, ensure_ascii=False, indent=4)


if __name__ == "__main__":
    logger.info("Parsing file: '%s'", DATA_PATH)
    logger.info("Using category: '%s'", DATA_PATH)
    try:
        with open(DATA_PATH, "r", encoding="utf8") as f:
            data = json.load(f)
        with open(CATEGORY_PATH, "r", encoding="utf8") as f:
            category = json.load(f)
        with open(ID_MAPPING_PATH, "r", encoding="utf8") as f:
            id_mapping = json.load(f)
    except FileNotFoundError:
        logger.critical("'%s' not found! Please parse the game bundles first.")
        sys.exit()
    i2languages = get_i2languages(data)

    parse_i18n()

    final_results = []
    parse_item_data(category, final_results)
    parse_recipe_data(data)

    with open(FINAL_RESULT_PATH, "w", encoding="utf8") as f:
        json.dump(final_results, f, ensure_ascii=False, indent=4)
